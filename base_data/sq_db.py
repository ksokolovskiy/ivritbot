import datetime
import sqlite3
import time
import random
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from contextlib import contextmanager

@contextmanager
def get_db_connection():
    con = sqlite3.connect("./base_data/db_bot.db")
    con.execute('pragma journal_mode=wal')
    try:
        yield con
        con.commit()
    except sqlite3.OperationalError as e:
        if 'database is locked' in str(e):
            time.sleep(0.1)  # небольшая задержка перед повторной попыткой
            yield con
            con.commit()
        else:
            raise
    finally:
        con.close()

def sql_create():
    with get_db_connection() as con:
        cur = con.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS users (
                    id TEXT PRIMARY KEY,
                    username TEXT,
                    access TEXT,
                    active TEXT,
                    last_time TEXT,
                    stop_keywords TEXT
                )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS verbs (
                    russian TEXT,
                    infin TEXT,
                    infin_man TEXT
                )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS time_step (
                        time_step TEXT,
                        last_send TEXT
                    )""")
        cur.close()

def add_user(id_user, username):
    with get_db_connection() as con:
        cur = con.cursor()
        now = datetime.datetime.now()
        cur.execute(f"INSERT INTO users VALUES(?,?,?,?,?,?)",(id_user, username, 'No', '60', now, ''))
        cur.close()

def get_verbs():
    with get_db_connection() as con:
        cur = con.cursor()
        list = cur.execute(f"SELECT * FROM verbs").fetchall()
        cur.close()
        return list

def get_users():
    with get_db_connection() as con:
        cur = con.cursor()
        list = cur.execute(f"SELECT * FROM users").fetchall()
        cur.close()
        return list

def active_user(user_id):
    with get_db_connection() as con:
        cur = con.cursor()
        cur.execute(f"UPDATE users SET active = 'Yes' WHERE id = ?", (user_id,))
        cur.close()

def add_stop_keyword(id_tg, keyword):
    with get_db_connection() as con:
        cur = con.cursor()
        list = cur.execute(f"SELECT stop_keywords FROM users WHERE id = ?", (id_tg,)).fetchall()
        list = list[0][0] + ' ' + keyword
        print(list)
        cur.execute(f"UPDATE users SET stop_keywords = ? WHERE id = ?", (list, id_tg,))
        cur.close()

def block_user(user_id):
    with get_db_connection() as con:
        cur = con.cursor()
        cur.execute(f"UPDATE users SET access = 'No' WHERE id = {user_id}")
        cur.close()

def white_user(user_id):
    with get_db_connection() as con:
        cur = con.cursor()
        cur.execute(f"UPDATE users SET access = 'Yes' WHERE id = ?", (user_id,))
        cur.close()

def add_verb(russian, infin, infin_man):
    with get_db_connection() as con:
        cur = con.cursor()
        cur.execute(f"INSERT INTO verbs VALUES(?,?,?)",(russian, infin, infin_man,))
        cur.close()

def get_excel():
    with get_db_connection() as con:
        df = pd.read_sql('select * from verbs', con)
        df.to_excel(r'result.xlsx', index=False)

def add_words():
    with get_db_connection() as con:
        df = pd.read_sql('select * from verbs', con)
        df.to_excel(r'files_xlsx/result.xlsx', index=False)
        fn = r"result.xlsx"
        wb = openpyxl.load_workbook(fn)
        ws = wb.active
        wb2 = openpyxl.load_workbook('files_xlsx/new_words.xlsx')
        ws2 = wb2.active
        new_words = []
        write_row = ws.max_row
        for row2 in range(1, ws2.max_row + 1):
            cell2 = ws2.cell(row=row2, column=1)
            if cell2.value:
                word_have = 0
                for row in range(1, ws.max_row + 1): # начало цикла проверки слова на наличие в существующей таблице
                    cell = ws.cell(row=row, column=1)
                    if cell.value == cell2.value:
                        word_have = 1
                if word_have == 0:
                    write_row = write_row + 1
                    ws[f'A{write_row}'] = cell2.value
                    ws[f'B{write_row}'] = ws2[f'B{row2}'].value
                    ws[f'C{write_row}'] = ws2[f'C{row2}'].value
        wb.save('files_xlsx/result.xlsx')
        save_db = pd.read_excel('files_xlsx/result.xlsx')
        save_db.to_sql(name='verbs', con=con, if_exists='replace', index=False)

def del_word(word):
    with get_db_connection() as con:
        df = pd.read_sql('select * from verbs', con)
        df.to_excel(r'files_xlsx/result.xlsx', index=False)
        fn = r"files_xlsx/result.xlsx"
        wb = openpyxl.load_workbook(fn)
        ws = wb.active
        status = False
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=1).value
            if cell == word:
                ws.delete_rows(row)
                status = True
                wb.save('files_xlsx/result.xlsx')
                save_db = pd.read_excel('files_xlsx/result.xlsx')
                save_db.to_sql(name='verbs', con=con, if_exists='replace', index=False)
        return status

def check_acess(username):
    with get_db_connection() as con:
        cur = con.cursor()
        list = cur.execute(f"SELECT access FROM users WHERE username = ?", (username,)).fetchall()
        print(list[0][0])
        access = list[0][0] == 'Yes'
        cur.close()
        return access

def get_timer():
    with get_db_connection() as con:
        cur = con.cursor()
        list = cur.execute(f"SELECT * FROM time_step").fetchall()
        cur.close()
        return list

def update_timer(now):
    with get_db_connection() as con:
        cur = con.cursor()
        cur.execute(f"UPDATE users SET last_time = ?", (now,))
        cur.close()

def update_time_step(step, id_tg):
    with get_db_connection() as con:
        cur = con.cursor()
        cur.execute(f"UPDATE users SET active = ? WHERE id = ?", (str(step), str(id_tg),))
        cur.close()

def update_time_step_for_user(step):
    with get_db_connection() as con:
        cur = con.cursor()
        cur.execute(f"UPDATE users SET active = ?", (step,))
        cur.close()

def add_white_user(username):
    with get_db_connection() as con:
        cur = con.cursor()
        now = datetime.datetime.now()
        cur.execute(f"INSERT INTO users VALUES(?,?,?,?,?,?)",('None', username, 'Yes', '60', now, ''))
        cur.close()